'''
import datetime
import openpyxl

wb = openpyxl.load_workbook('2022年股票数据.xlsx')

# 获取工作表的名字
print(wb.sheetnames)
sheet=wb.worksheets[0]
print(sheet.dimensions)
print(sheet.max_row,sheet.max_column)

# 获取指定单元格的值
print(sheet.cell(3,3).value)
print(sheet['C3'].value)
print(sheet['G255'].value)

# 获取多个单元格（嵌套元组）
print(sheet['A2:C5'])

#读取所有单元格的数据
i=1
for row_ch in range(1, sheet.max_row + 1):
    i=i+1
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'),end='\t')
        elif type(value) == int:
            print(f'{value:<10d}',end='\t')
        elif type(value) == float:
            print(f'{value:.4f}',end='\t')
        else:
            print(value,end='\t')
    if i>10:
        break
    print()
'''

'''
#写Excel文件

import random
import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = '期末成绩'

titles = ('姓名','语文', '数学', '英语')
for col_index,title in enumerate(titles):
    sheet.cell(1, col_index+1,title)

names = ('关羽', '张飞', '赵云', '马超', '黄忠')
for row_index,name in enumerate(names):
    sheet.cell(row_index + 2,1,name)
    for col_index in range(2,5):
        sheet.cell(row_index +2, col_index, random.randrange(50,101))

wb.save('resources/考试成绩表.xlsx')

'''

'''
#调整样式和公式计算
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# 对齐方式
alignment = Alignment(horizontal='center', vertical='center')
# 边框线条
side = Side(color='ff7f50', style='mediumDashed')

wb = openpyxl.load_workbook('resources/考试成绩表.xlsx')
sheet = wb.worksheets[0]

# 调整行高和列宽
sheet.row_dimensions[1].height = 30
sheet.column_dimensions['E'].width = 120

sheet['E1'] = '平均分'
# 设置字体
sheet.cell(1, 5).font = Font(size=18, bold=True, color='ff1493', name='华文楷体')
# 设置对齐方式
sheet.cell(1, 5).alignment = alignment
# 设置单元格边框
sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
for i in range(2, 7):
    # 公式计算每个学生的平均分
    sheet[f'E{i}'] = f'=average(B{i}:D{i})'
    sheet.cell(i, 5).font = Font(size=12, color='4169e1', italic=True)
    sheet.cell(i, 5).alignment = alignment

wb.save('resources/考试成绩表.xlsx')
'''


#生成统计图表

from openpyxl import Workbook
from openpyxl.chart import BarChart,Reference

wb = Workbook(write_only=True)
sheet = wb.create_sheet()

rows = [
    ('类别','销售A组','销售B组'),
    ('手机',40,30),
    ('平板',50,60),
    ('笔记本',80,70),
    ('外围设备',20,10),
]

for row in rows:
    sheet.append(row)


chart = BarChart()
chart.type = 'col'
chart.style = 10

chart.title = '销售统计图'

chart.y_axis.title = '销量'

chart.x_axis.title = '商品类别'

data = Reference(sheet,min_col=2,min_row=1,max_row=5,max_col=3)

cats = Reference(sheet,min_col=1,min_row=2,max_row=5)

chart.add_data(data, titles_from_data=True)

chart.set_categories(cats)
chart.shape = 4

sheet.add_chart(chart,'A10')

wb.save('resources/demo.xlsx')



















